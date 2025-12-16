from openpyxl import load_workbook
from Models.hotelExcel import HotelExcel, PeriodoGroup, DatosExcel
from Models.periodo import Periodo
from ExtractorDatos.utils import *
import traceback
import sys

EXCLUSIONES = [
    "season rates", "(per room)", "closing agreement",
    "rates includes", "promotion", "not included", "high speed", "minimum stay",
    "other benefits"
]

LEYENDAS_AGREEMENT = ["closing agreement"]

def cargar_excel(path_excel, max_row=300) -> DatosExcel:
    """Carga datos de hoteles, habitaciones y periodos desde Excel.

    Args:
        path_excel: Ruta al archivo Excel
        max_row: Número máximo de filas a procesar (default: 300)

    Returns:
        DatosExcel con lista de hoteles extraídos

    La función ha sido refactorizada usando ContextoExtraccion para mejorar
    la mantenibilidad y testabilidad del código. La lógica de procesamiento
    se divide en métodos especializados para cada tipo de entidad.
    """
    from ExtractorDatos.contexto_extraccion import ContextoExtraccion

    wb = load_workbook(path_excel)
    ws = wb.active

    # Crear contexto de extracción
    ctx = ContextoExtraccion(ws=ws)

    # Procesar cada fila
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_row)):  # type: ignore
        ctx.procesar_fila(row, i)

    return DatosExcel(hoteles=ctx.hoteles)


def detectar_y_agregar_periodo(hotel_actual: HotelExcel, periodo_raw, hubo_nombre_periodo: bool, i) -> tuple[bool, str | None]:
    """
    Detecta y agrega periodos al hotel.

    Returns:
        tuple: (continuar: bool, nuevo_nombre_periodo: str | None)
            - continuar: True si se debe saltar esta fila sin procesar habitaciones
            - nuevo_nombre_periodo: nombre del periodo detectado (solo si contiene 'season'), None en caso contrario
    """
    nuevo_nombre = None
    periodo_raw = str(periodo_raw)
    periodo_sin_nombre_creado = False

    # Ignorar URLs y otros formatos no válidos
    if isinstance(periodo_raw, str) and ("http" in periodo_raw.lower() or "https" in periodo_raw.lower()):
        return (True, None)

    ## Se intentan detectar nombres de periodos tipo "low season", "high season", SPECIAL DATES
    if contiene_season(periodo_raw):
        nuevo_nombre = str(periodo_raw).strip()
        agregar_nombre_group(hotel_actual, nuevo_nombre)
        return (False, nuevo_nombre)
    
    if contiene_special_dates(periodo_raw):
        nuevo_nombre = str(periodo_raw).strip()
        agregar_nombre_group(hotel_actual, nuevo_nombre)
        return (False, nuevo_nombre)

    ## Si no hay nombre de periodo, se intenta EXTRAER FECHAS
    try:
        ## intentamos extraer las fechas con parentesis (1May25 - 30Sep25) (12Dec25 - 27Dec25)
        fechas_con_parentesis = extraer_fechas_con_parentesis(periodo_raw)
        
        if fechas_con_parentesis:
            # IMPORTANTE: Iterar sobre TODOS los periodos encontrados
            for fecha_inicio, fecha_fin in fechas_con_parentesis:
                nombre_periodo_a_usar = "SIN NOMBRE DE GRUPO"
                # Construir el periodo
                periodo = construir_periodo(fecha_inicio, fecha_fin)

                # Agregarlo al hotel

                if hubo_nombre_periodo:
                    hotel_actual.periodos_group[-1].periodos.append(periodo)
                else:
                    if len(fechas_con_parentesis) > 1:
                        if not periodo_sin_nombre_creado:
                            periodo_group = PeriodoGroup(
                            nombre=nombre_periodo_a_usar,
                            periodos=[periodo]
                        )
                            hotel_actual.periodos_group.append(periodo_group)
                            periodo_sin_nombre_creado = True
                        else: 
                            hotel_actual.periodos_group[-1].periodos.append(periodo)
                
                    else:
                        if not periodo_sin_nombre_creado:
                            periodo_group = PeriodoGroup(
                            nombre=nombre_periodo_a_usar,
                            periodos=[periodo]
                        )
                            hotel_actual.periodos_group.append(periodo_group)
                            periodo_sin_nombre_creado = True
                        else:
                            hotel_actual.periodos_group[-1].periodos.append(periodo)
    
            if nombre_periodo_a_usar == "SIN NOMBRE DE GRUPO":
                return (False, nombre_periodo_a_usar)
            else:
                return (False, None)

        else: ## caso SPECIAL DATES 
            ## Sino, intentamos extraer fechas del tipo "New Year: 26Dec25 - 3Jan26" "Easter: 2-5Apr26"
            resultado = extraer_fechas_sin_parentesis(periodo_raw)

            ## Si no se pudo extraer nada, se salta
            if not resultado:
                return (True, None)

            nombre_periodo_extraido, fechas_sin_parentesis = resultado

            ## Imposible extraer fechas
            if not fechas_sin_parentesis:
                return (True, None)

            ## se construye el periodo y se lo agrega al hotel actual
            fecha_inicio, fecha_fin = fechas_sin_parentesis


            periodo = construir_periodo(fecha_inicio, fecha_fin, nombre_periodo_extraido)
            hotel_actual.periodos_group[-1].periodos.append(periodo)

            return (False, None)

    except Exception as e:
        _, _, exc_traceback = sys.exc_info()
        print(f"[ERROR] Fila {i}:")
        print(f"  - Dato original: {periodo_raw}")
        print(f"  - Tipo de error: {type(e).__name__}")
        print(f"  - Mensaje: {str(e)}")
        print(f"  - Hotel actual: {hotel_actual.nombre if hotel_actual else 'None'}")
        print("  - Traceback:")
        for filename, lineno, funcname, line in traceback.extract_tb(exc_traceback):
            print(f"    Archivo: {filename}")
            print(f"    Línea: {lineno}")
            print(f"    Función: {funcname}")
            if line:
                print(f"    Código: {line}")
        print("  -------------------")
        return (True, None)  # En caso de error, continuar a la siguiente fila
