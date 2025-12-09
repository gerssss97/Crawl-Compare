from openpyxl import load_workbook
from Models.hotelExcel import *
from Models.periodo import *
from ExtractorDatos.utils import *
import traceback
import sys


EXCLUSIONES = [
    "season rates", "(per room)", "closing agreement",
    "rates includes", "promotion", "not included", "high speed", "minimum stay",
    "other benefits"
]

LEYENDAS_AGREEMENT = ["closing agreement"]

def cargar_excel(path_excel, max_row=300) -> DatosExcel:
    wb = load_workbook(path_excel)
    ws = wb.active

    hoteles: list[HotelExcel] = []
    hotel_actual: HotelExcel | None = None
    tipo_actual: TipoHabitacionExcel | None = None

    precio_str = None
    fila_nombre_seasson = -1
    nombre_periodo = None 

    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_row)):  # type: ignore
        if not any(row):
            continue  # fila vacía

        ## Logica para extraer periodos
        periodo_raw = row[1]

        # Solo procesar si parece un periodo válido
        if periodo_raw is not None and parece_periodo(periodo_raw):
            continuar, nuevo_nombre = detectar_y_agregar_periodo(hotel_actual, periodo_raw, i, fila_nombre_seasson, nombre_periodo)

            # Si se detectó un nuevo nombre de periodo, persistirlo
            if nuevo_nombre is not None:
                nombre_periodo = nuevo_nombre
                fila_nombre_seasson = i

            if continuar:
                # Si es un nombre de periodo (no fechas), no hacemos nada aquí
                continue

        if row[0] is None:
            continue  # sin nombre

        nombre_raw = str(row[0]).strip()
        nombre_norm = nombre_raw.lower()

        ##  Saltar exclusiones
        if any(nombre_norm.startswith(excl) for excl in EXCLUSIONES):
            continue

        ##  Detectar NUEVO HOTEL (empieza con "hotel" o patrón similar)
        if nombre_raw.endswith("(A)"):
            # Resetear variables de periodo para el nuevo hotel
            fila_nombre_seasson = -1
            nombre_periodo = None

            hotel_actual = HotelExcel(nombre=nombre_raw, tipos=[], habitaciones_directas=[],periodos=[])
            hoteles.append(hotel_actual)
            primer_habitacion = True
            tipo_actual = None
            continue

        ## Detectar TIPO de habitación (mayúsculas + sin precio en la fila)
        if nombre_raw.isupper() and row[2] is None:
            nombre_raw = nombre_raw.replace(":", "").strip()
            tipo_actual = TipoHabitacionExcel(nombre=nombre_raw, habitaciones=[])
            if hotel_actual:
                hotel_actual.tipos.append(tipo_actual)
            continue

        ## Logica de los PRECIOS de las habitaciones
        col_precio = obtener_valor_real(ws, i, 2)  # columna 2 = índice C en Excel
        valor_str = str(col_precio).strip().lower() if col_precio is not None else ""

        es_leyenda_agreement = valor_str in LEYENDAS_AGREEMENT
        es_numerico = valor_str.replace(".", "").replace(",", "").isdigit()
        es_texto_no_leyenda = valor_str and not es_numerico and not es_leyenda_agreement
        ## Lógica para detectar "closing agreement" o similares, omitiendo "habitaciones" cuyas leyendas NO esten en la lista
        ## ademas de reiniciar el precio_str

        if es_texto_no_leyenda:
            continue # Ejemplo: "closing"  → ignorar la habitación


        precio = None
        if es_leyenda_agreement:
            precio_str = valor_str
            precio = valor_str
        elif es_numerico:
            precio_str = None
            precio = valor_str
        elif precio_str:
            precio = precio_str


        es_habitacion = nombre_norm.startswith(("dbl", "sgl", "tpl"))
        # chequeo si es una habitacion (dbl, sgl, tpl), si no tiene precio chequeo precio BAR
        if (es_habitacion):
            if not precio and row[3] is not None:
                precio = str(row[3]).strip()

            ## si sigue sin precio luego de intentar con precio BAR, se omite
            if not precio:
                continue

            habitacion = HabitacionExcel(
                nombre=nombre_raw,
                precio=precio,
                row_idx=i,
                periodo_ids=[]  # se asignan a continuación
            )

            # Asignar periodos a la habitación
            if hotel_actual and hotel_actual.periodos_group:
                # Siempre asignar el último grupo de periodos disponible
                ultimo_grupo = hotel_actual.periodos_group[-1]
                for periodo in ultimo_grupo.periodos:
                    habitacion.periodo_ids.append(periodo.id)

            if hotel_actual:
                if tipo_actual:
                    tipo_actual.habitaciones.append(habitacion)
                else:
                    hotel_actual.habitaciones_directas.append(habitacion)
        else:
            if not precio:
                continue
            # si no es habitacion y no tiene precio, es un extra
            extra = Extra(
                nombre=nombre_raw,
                precio=precio
            )
            if hotel_actual:
                hotel_actual.extras.append(extra)

    ## agregar periodos a las habitaciones del ultimo hotel
    # agregar_periodos_a_habitaciones(hotel_actual)
    return DatosExcel(hoteles=hoteles)


def detectar_y_agregar_periodo(hotel_actual: HotelExcel, periodo_raw, i, fila_nombre_seasson, nombre_periodo_actual):
    """
    Detecta y agrega periodos al hotel.

    Returns:
        tuple: (continuar: bool, nuevo_nombre_periodo: str | None)
            - continuar: True si se debe saltar esta fila sin procesar habitaciones
            - nuevo_nombre_periodo: nombre del periodo detectado (solo si contiene 'season'), None en caso contrario
    """
    nuevo_nombre = None
    periodo_raw = str(periodo_raw)

    # Ignorar URLs y otros formatos no válidos
    if isinstance(periodo_raw, str) and ("http" in periodo_raw.lower() or "https" in periodo_raw.lower()):
        return (True, None)

    ## Se intentan detectar nombres de periodos tipo "low season", "high season", etc
    if contiene_season(periodo_raw):
        nuevo_nombre = str(periodo_raw).strip()
        # Retornar sin crear periodo, solo guardamos el nombre para los siguientes
        return (True, nuevo_nombre)

    ## Si no hay nombre de periodo, se intenta extraer fechas
    try:
        ## intentamos extraer las fechas con parentesis (1May25 - 30Sep25) (12Dec25 - 27Dec25)
        fechas_con_parentesis = extraer_fechas_con_parentesis(periodo_raw)

        if fechas_con_parentesis:
            # IMPORTANTE: Iterar sobre TODOS los periodos encontrados
            for fecha_inicio, fecha_fin in fechas_con_parentesis:
                ## El nombre del periodo se determina por la distancia
                ## respecto a la fila donde se encontro una leyenda con "season"
                if fila_nombre_seasson != -1 and i - fila_nombre_seasson <= 3 and nombre_periodo_actual:
                    nombre_periodo_a_usar = nombre_periodo_actual
                    # Agregar a un grupo existente con el mismo nombre
                    agregar_periodo_a_grupo_existente = True
                else:
                    nombre_periodo_a_usar = "hardoceado con parentesis"
                    agregar_periodo_a_grupo_existente = False

                # Construir el periodo
                periodo = construir_periodo(fecha_inicio, fecha_fin)

                # Agregarlo al hotel
                if agregar_periodo_a_grupo_existente and hotel_actual.periodos_group:
                    # Buscar el grupo con el nombre correcto
                    grupo_encontrado = False
                    for grupo in hotel_actual.periodos_group:
                        if grupo.nombre == nombre_periodo_a_usar:
                            grupo.periodos.append(periodo)
                            grupo_encontrado = True
                            break

                    # Si no existe el grupo, crearlo
                    if not grupo_encontrado:
                        periodo_group = PeriodoGroup(
                            nombre=nombre_periodo_a_usar,
                            periodos=[periodo]
                        )
                        hotel_actual.periodos_group.append(periodo_group)
                else:
                    # Crear nuevo grupo
                    periodo_group = PeriodoGroup(
                        nombre=nombre_periodo_a_usar,
                        periodos=[periodo]
                    )
                    hotel_actual.periodos_group.append(periodo_group)

            return (False, None)  # No continuar, procesamos fechas exitosamente

        else:
            ## Sino, intentamos extraer fechas del tipo "New Year: 26Dec25 - 3Jan26" "Easter: 2-5Apr26"
            resultado = extraer_fechas_sin_parentesis(periodo_raw)

            ## Si no se pudo extraer nada, se salta
            if not resultado:
                return (True, None)

            nombre_periodo_extraido, fechas_sin_parentesis = resultado

            ## Imposible extraer fechas
            if not fechas_sin_parentesis:
                return (True, None)

            ## se construye el periodo y se lo agrega al hotel actual
            fecha_inicio, fecha_fin = fechas_sin_parentesis

            if not nombre_periodo_extraido:
                nombre_periodo_a_usar = "hardoceado sin parentesis"
            else:
                nombre_periodo_a_usar = nombre_periodo_extraido

            periodo = construir_periodo(fecha_inicio, fecha_fin)

            # Crear nuevo grupo para periodos sin paréntesis
            periodo_group = PeriodoGroup(
                nombre=nombre_periodo_a_usar,
                periodos=[periodo]
            )
            hotel_actual.periodos_group.append(periodo_group)

            return (False, None)

    except Exception as e:
        _, _, exc_traceback = sys.exc_info()
        print(f"[ERROR] Fila {i}:")
        print(f"  - Dato original: {periodo_raw}")
        print(f"  - Tipo de error: {type(e).__name__}")
        print(f"  - Mensaje: {str(e)}")
        print(f"  - Hotel actual: {hotel_actual.nombre if hotel_actual else 'None'}")
        print("  - Traceback:")
        for filename, lineno, funcname, line in traceback.extract_tb(exc_traceback):
            print(f"    Archivo: {filename}")
            print(f"    Línea: {lineno}")
            print(f"    Función: {funcname}")
            if line:
                print(f"    Código: {line}")
        print("  -------------------")
        return (True, None)  # En caso de error, continuar a la siguiente fila
