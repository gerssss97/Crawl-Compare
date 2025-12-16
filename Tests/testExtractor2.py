"""
Script de Testing del Extractor de Datos Excel
------------------------------------------------
Este script carga datos de un Excel usando el extractor,
y exporta TODA la información extraída a un nuevo archivo Excel
para validación visual y comparación.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ExtractorDatos.extractor import cargar_excel


def crear_hoja_periodos(wb, hotel):
    """
    Crea una hoja con todos los periodos de un hotel.

    Args:
        wb: Workbook de openpyxl
        hotel: Objeto HotelExcel
    """
    nombre_hoja = f"Periodos_{hotel.nombre[:20]}"  # Limitar longitud del nombre
    ws = wb.create_sheet(nombre_hoja)

    # Encabezados
    headers = ["ID", "Nombre", "Fecha Inicio", "Fecha Fin"]
    ws.append(headers)

    # Estilo para encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos de periodos
    for periodoGroup in hotel.periodos_group:
        for periodo in periodoGroup.periodos:
            ws.append([
                periodo.id,
                periodoGroup.nombre,
                periodo.fecha_inicio.strftime("%d-%m-%Y"),
                periodo.fecha_fin.strftime("%d-%m-%Y")
            ])

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    return ws


def crear_hoja_extras(wb, hotel):
    """
    Crea una hoja con todos los extras de un hotel.

    Args:
        wb: Workbook de openpyxl
        hotel: Objeto HotelExcel
    """
    nombre_hoja = f"Extras_{hotel.nombre[:20]}"  # Limitar longitud del nombre
    ws = wb.create_sheet(nombre_hoja)

    # Encabezados
    headers = ["Hotel", "Nombre del Extra", "Precio"]
    ws.append(headers)

    # Estilo para encabezados
    header_fill = PatternFill(start_color="9966FF", end_color="9966FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos de extras
    if hotel.extras:
        for extra in hotel.extras:
            precio = extra.precio if extra.precio is not None else ""
            ws.append([
                hotel.nombre,
                extra.nombre,
                precio
            ])
    else:
        ws.append([hotel.nombre, "(Sin extras registrados)", ""])

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15

    return ws


def crear_hoja_tipos(wb, hotel):
    """
    Crea una hoja con todos los tipos de habitación de un hotel.

    Args:
        wb: Workbook de openpyxl
        hotel: Objeto HotelExcel
    """
    nombre_hoja = f"Tipos_{hotel.nombre[:20]}"  # Limitar longitud del nombre
    ws = wb.create_sheet(nombre_hoja)

    # Encabezados
    headers = ["Tipo de Habitación", "Cantidad de Habitaciones", "Nombres de Habitaciones"]
    ws.append(headers)

    # Estilo para encabezados
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar información de cada tipo
    if hotel.tipos:
        for tipo in hotel.tipos:
            cantidad_hab = len(tipo.habitaciones)

            if cantidad_hab > 0:
                nombres_hab = ", ".join([hab.nombre for hab in tipo.habitaciones])
            else:
                nombres_hab = "(Sin habitaciones asociadas)"

            ws.append([
                tipo.nombre,
                cantidad_hab,
                nombres_hab
            ])
    else:
        ws.append(["(No hay tipos de habitación definidos)", 0, ""])

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 80

    return ws


def obtener_detalles_periodos(hotel, periodo_ids):
    """
    Convierte una lista de IDs de periodo en una cadena legible con los detalles.

    Args:
        hotel: Objeto HotelExcel
        periodo_ids: Lista de IDs de periodo

    Returns:
        String con detalles de los periodos
    """
    if not periodo_ids:
        return "Sin periodos asignados"

    detalles = []
    for pid in periodo_ids:
        periodo = hotel.periodo_por_id(pid)
        if periodo:
            detalles.append(f" ({periodo.fecha_inicio.strftime('%d-%m-%Y')} - {periodo.fecha_fin.strftime('%d-%m-%Y')})")
        else:
            detalles.append(f"ID {pid} (no encontrado)")

    return " | ".join(detalles)


def crear_hoja_habitaciones(wb, hotel):
    """
    Crea una hoja con todas las habitaciones de un hotel (directas y por tipo).

    Args:
        wb: Workbook de openpyxl
        hotel: Objeto HotelExcel
    """
    nombre_hoja = f"Resumen_{hotel.nombre[:20]}"  # Limitar longitud del nombre
    ws = wb.create_sheet(nombre_hoja)

    # Encabezados
    headers = [
        "Hotel",
        "Tipo",
        "Habitación",
        "Precio",
        "Precio String",
        "Row Index",
        "Periodo IDs",
        "Detalles de Periodos"
    ]
    ws.append(headers)

    # Estilo para encabezados
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar habitaciones directas (sin tipo específico)
    for habitacion in hotel.habitaciones_directas:
        precio_num = habitacion.precio if isinstance(habitacion.precio, (int, float)) else ""
        precio_str = habitacion.precio_string if habitacion.precio_string else ""
        periodo_ids_str = ", ".join(map(str, habitacion.periodo_ids)) if habitacion.periodo_ids else ""
        detalles_periodos = obtener_detalles_periodos(hotel, habitacion.periodo_ids)

        ws.append([
            hotel.nombre,
            "(Sin tipo específico)",
            habitacion.nombre,
            precio_num,
            precio_str,
            habitacion.row_idx,
            periodo_ids_str,
            detalles_periodos
        ])

    # Agregar habitaciones organizadas por tipo
    for tipo in hotel.tipos:
        for habitacion in tipo.habitaciones:
            precio_num = habitacion.precio if isinstance(habitacion.precio, (int, float)) else ""
            precio_str = habitacion.precio_string if habitacion.precio_string else ""
            periodo_ids_str = ", ".join(map(str, habitacion.periodo_ids)) if habitacion.periodo_ids else ""
            detalles_periodos = obtener_detalles_periodos(hotel, habitacion.periodo_ids)

            ws.append([
                hotel.nombre,
                tipo.nombre,
                habitacion.nombre,
                precio_num,
                precio_str,
                habitacion.row_idx,
                periodo_ids_str,
                detalles_periodos
            ])

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 60

    return ws


def exportar_datos_a_excel(datos_excel, archivo_salida):
    """
    Exporta todos los datos extraídos a un archivo Excel para validación.

    Args:
        datos_excel: Objeto DatosExcel con todos los hoteles
        archivo_salida: Ruta del archivo Excel de salida
    """
    wb = Workbook()

    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Crear hoja de resumen general
    ws_resumen = wb.create_sheet("Resumen General", 0)
    ws_resumen.append(["Información General de la Extracción"])
    ws_resumen.append([])
    ws_resumen.append(["Total de Hoteles:", len(datos_excel.hoteles)])
    ws_resumen.append([])
    ws_resumen.append(["Hotel", "Total Periodos", "Total Tipos", "Total Habitaciones Directas", "Total Habitaciones por Tipo", "Total Extras"])

    # Estilo para título
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Estilo para encabezados de tabla
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws_resumen[5]:
        cell.fill = header_fill
        cell.font = header_font

    # Agregar información de cada hotel
    for hotel in datos_excel.hoteles:
        total_hab_por_tipo = sum(len(tipo.habitaciones) for tipo in hotel.tipos)
        ws_resumen.append([
            hotel.nombre,
            len(hotel.periodos_group),
            len(hotel.tipos),
            len(hotel.habitaciones_directas),
            total_hab_por_tipo,
            len(hotel.extras)
        ])

    # Ajustar anchos de columnas en resumen
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 18
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 28
    ws_resumen.column_dimensions['E'].width = 30
    ws_resumen.column_dimensions['F'].width = 15

    # Crear hojas para cada hotel
    for hotel in datos_excel.hoteles:
        print(f"Exportando datos del hotel: {hotel.nombre}")

        # Crear hoja de habitaciones
        crear_hoja_habitaciones(wb, hotel)

        # Crear hoja de tipos de habitación
        # crear_hoja_tipos(wb, hotel)  # Deshabilitado - función disponible si se necesita

        # Crear hoja de periodos
        crear_hoja_periodos(wb, hotel)

        # Crear hoja de extras
        # crear_hoja_extras(wb, hotel)  # Deshabilitado - función disponible si se necesita

    # Guardar archivo
    wb.save(archivo_salida)
    print(f"\n[OK] Archivo exportado exitosamente: {archivo_salida}")
    print(f"     Total de hojas creadas: {len(wb.sheetnames)}")


def main():
    """
    Función principal que ejecuta el proceso de extracción y exportación.
    """
    print("=" * 60)
    print("TESTING DEL EXTRACTOR DE DATOS - EXCEL")
    print("=" * 60)
    print()

    # Configuración de rutas
    path_excel_entrada = "./Data/Extracto_prueba2.xlsx"
    path_excel_salida = "./Data/Extracto_Validacion2.xlsx"

    print(f"[>] Archivo de entrada: {path_excel_entrada}")
    print(f"[>] Archivo de salida:  {path_excel_salida}")
    print()

    # Cargar datos del Excel
    print("[*] Cargando y extrayendo datos del Excel...")
    datos = cargar_excel(path_excel_entrada)

    print(f"[OK] Extracción completada")
    print(f"     - Total de hoteles extraídos: {len(datos.hoteles)}")
    print()

    # Mostrar resumen de lo extraído
    print("RESUMEN DE DATOS EXTRAIDOS:")
    print("-" * 60)
    for hotel in datos.hoteles:
        total_hab_por_tipo = sum(len(tipo.habitaciones) for tipo in hotel.tipos)
        total_habitaciones = len(hotel.habitaciones_directas) + total_hab_por_tipo

        print(f"\nHOTEL: {hotel.nombre}")
        print(f"   |-- Periodos_ group: {len(hotel.periodos_group)}")
        print(f"   |-- Tipos de habitación: {len(hotel.tipos)}")
        print(f"   |-- Habitaciones directas: {len(hotel.habitaciones_directas)}")
        print(f"   |-- Habitaciones por tipo: {total_hab_por_tipo}")
        print(f"   |-- TOTAL HABITACIONES: {total_habitaciones}")
        print(f"   |-- Extras: {len(hotel.extras)}")

    print()
    print("-" * 60)
    print()

    # Exportar a Excel
    print("[*] Exportando datos a Excel para validación...")
    exportar_datos_a_excel(datos, path_excel_salida)

    print()
    print("=" * 60)
    print("PROCESO COMPLETADO CON EXITO")
    print("=" * 60)
    print()
    print(">> Abre el archivo para validar visualmente los datos extraídos:")
    print(f"   {path_excel_salida}")
    print()


if __name__ == "__main__":
    main()
